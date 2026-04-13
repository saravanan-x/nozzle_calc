from openpyxl import Workbook
import numpy as np
import math
import os

def result(f, m, cd_mm, lc_mm, conv_mm, dt_mm, div_mm, de_mm, total_mm,
    ar, ρ0, vt, ρt, me, mg, mc, ep, et, v_e, isp, m0, F):


    wb = Workbook()
    ws = wb.active
    ws.title = "Nozzle Report"

    # ------------------ TITLE ------------------
    ws["B1"] = "ROCKET NOZZLE DESIGN RESULT"

    # ------------------ GEOMETRY (LEFT SIDE) ------------------
    ws["A3"] = "ROCKET NOZZLE GEOMETRY"
    
    ws["A5"] = f"Required Thrust       : {f} N"
    ws["A6"] = f"Required Mass flow    : {m:.4f} kg/s"
    ws["A7"] = f"Chamber Diameter      : {cd_mm:.2f} mm"
    ws["A8"] = f"Chamber Length        : {lc_mm:.2f} mm"
    ws["A9"] = f"Convergent Length     : {conv_mm:.2f} mm"
    ws["A10"] = f"Throat Diameter       : {dt_mm:.2f} mm"
    ws["A11"] = f"Divergent Length      : {div_mm:.2f} mm"
    ws["A12"] = f"Exit Diameter         : {de_mm:.2f} mm"
    ws["A13"] = f"Total Nozzle Length   : {total_mm:.2f} mm"

    # ------------------ FLOW (RIGHT SIDE) ------------------
    ws["c3"] = "ROCKET NOZZLE FLOW PROPERTIES"

    ws["C5"] = f"Nozzle Area Ratio     : {ar:.4f}"
    ws["C6"] = f"Stagnation Density    : {ρ0:.4f} kg/m³"
    ws["C7"] = f"Throat Velocity       : {vt:.4f} m/s"
    ws["C8"] = f"Throat Density        : {ρt:.4f} kg/m³"
    ws["C9"] = f"Exit Mach number      : {me:.4f}"
    ws["C10"] = f"General mass flow     : {mg:.4f} kg/s"
    ws["C11"] = f"Choked mass flow      : {mc:.4f} kg/s"
    ws["C12"] = f"Exit Pressure         : {ep:.4f} Pa"
    ws["C13"] = f"Exit Temperature      : {et:.4f} K"
    ws["C14"] = f"Exit Velocity         : {v_e:.4f} m/s"
    ws["C15"] = f"Specific Impulse      : {isp:.4f} S"
    ws["C16"] = f"Exit Mass flow rate   : {m0:.4f} kg/s"
    ws["C17"] = f"Final Thrust          : {F:.4f} N"

    # ------------------ SAVE ------------------
    save_path = "/home/saravanan/Music/nozzle.py/result"
    os.makedirs(save_path, exist_ok=True)

    file_path = os.path.join(save_path, "nozzle_report.xlsx")

    wb.save(file_path)